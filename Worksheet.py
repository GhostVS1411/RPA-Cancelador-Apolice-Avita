from openpyxl import Workbook, load_workbook
from os.path import exists
from os import listdir
from unidecode import unidecode

class Worksheet:

    def __init__(self,caminho):
        self.caminho = caminho
        
    def create(self):
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def read(self):
        try:
            self.wb = load_workbook(self.caminho)
            self.ws = self.wb.active
        except:
            print("A planilha não existe ou o caminho está incorreto")
        
##    def format(self,colA='',colB='',colC='',colD='',colE='',width = 30):
##        self.ws['A1'] = colA
##        self.ws['B1'] = colB
##        self.ws['C1'] = colC
##        self.ws['D1'] = colD
##        self.ws['E1'] = colE
##       #ou
##        for coluna in ['A','B','C','D','E']:
##            self.ws.column_dimensions[coluna].width = width
##            self.ws[coluna + "1"] = [colA,colB,colC,colD,colE][['A','B','C','D','E'].index(coluna)]

    def add(self,colA='',colB='',colC='',colD='',colE=''):
        self.ws.append([colA,colB,colC,colD,colE])

    def save(self):
        try:
            self.wb.save(self.caminho)
        except:
            print("A planilha está aberta")
            print("Feche para poder salvar")

def collectData(numApolice,reclamantes,processo,rteSemPasta,apSemPasta,procSemPasta,rteSemArquivo,ws_dados,addRelatorio,saveRelatorio,caminho):

    charReclamantes = ''
    charProcessos = ''
    charApolice = ''

    for column in range(1,ws_dados.max_column):

        cell = (ws_dados[chr(column+64)+'1'].value)

        if cell != None:
            
            if 'reclamante' in cell.lower():
                charReclamantes = chr(column+64)
                print(f"Reclamantes: coluna {reclamantes}")
            elif 'processo' in cell.lower():
                charProcessos = chr(column+64)
                print(f"Processos: coluna {processos}")
            elif 'apólice' in cell.lower():
                charApolice = chr(column+64)
                print(f"Apólices coluna {apolice}")

    if (charReclamantes == ''):
        print("Falta a coluna dos remetentes")
    elif (charProcessos == ''):
        print("Falta a coluna dos processos")
    elif (charApolice == ''):
        print("Falta a coluna da apólice")
    else:       
        for linha in range(2,ws_dados.max_row+1):

            pasta = caminho.replace("pasta",unidecode(ws_dados[charReclamantes + str(linha)].value))
        
            if exists(pasta) and (listdir(pasta) != []) :
           
                numApolice.append(ws_dados[charApolice + str(linha)].value)
                reclamantes.append(unidecode(ws_dados[charReclamantes + str(linha)].value))
                processo.append(ws_dados[charProcessos + str(linha)].value)
                
            elif (exists(pasta) == False):
                rteSemPasta.append(unidecode(ws_dados[charReclamantes + str(linha)].value))
                apSemPasta.append(ws_dados[charApolice + str(linha)].value)
                procSemPasta.append(ws_dados[charProcessos + str(linha)].value)

                addRelatorio(procSemPasta[-1],rteSemPasta[-1],'',apSemPasta[-1],'Pasta não encontrada')
                saveRelatorio() 
            else:
                rteSemArquivo.append(unidecode(ws_dados[charReclamantes + str(linha)].value))

                addRelatorio(ws_dados[charProcessos + str(linha)].value,rteSemArquivo[-1],'',ws_dados['H' + str(linha)].value,'Pasta Sem Arquivos')
                saveRelatorio()
##
            
##relatorio.format("teste","da","planilha","esse","é")
##relatorio.add("a","b","c","d","e")
##relatorio.save()
## 
##baseDados = Worksheet("baseDados.xlsx")
##baseDados.read()
##baseDados.add(1,2,3,4,5)
##baseDados.save()

#collectData(numApolice,reclamantes,processo,rteSemPasta,apSemPasta,procSemPasta,rteSemArquivo,baseDados.ws,relatorio.add,relatorio.save,caminho)










